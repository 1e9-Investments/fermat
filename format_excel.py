#!/usr/bin/env python3
"""
Fermat Excel Formatter - Applies 1e9-Investments Excel standards
Full spec: row heights, fonts, colors, navigation, charts, freeze panes, etc.
"""

import sys
import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from copy import copy

# === COLOR PALETTE (Bright Pastels) ===
PASTEL_GREEN = 'C6EFCE'
PASTEL_YELLOW = 'FFE699'
PASTEL_BLUE = 'DEEBF7'
PASTEL_ORANGE = 'FCE4D6'

# === FONT SETTINGS - 11pt Aptos ===
DEFAULT_FONT = Font(name='Aptos', size=11, color='000000')
HEADER_FONT = Font(name='Aptos', size=11, color='000000', bold=True)
TITLE_FONT = Font(name='Aptos', size=12, color='000000', bold=True)
NAV_FONT = Font(name='Wingdings', size=11, color='000000')
LABEL_FONT = Font(name='Aptos', size=11, color='000000')

# === WALL STREET COLOR CODING (Data Source Convention) ===
# Blue: Hardcoded values / imported data
FONT_HARDCODE = Font(name='Aptos', size=11, color='0000FF')
# Black: Formulas / calculated values
FONT_FORMULA = Font(name='Aptos', size=11, color='000000')
# Green: References to other sheets in same workbook
FONT_SHEET_REF = Font(name='Aptos', size=11, color='006400')
# Purple: References to other workbooks
FONT_WORKBOOK_REF = Font(name='Aptos', size=11, color='800080')

# === FILLS - Bright pastels only, NO dark fills ===
GREEN_FILL = PatternFill(start_color=PASTEL_GREEN, end_color=PASTEL_GREEN, fill_type='solid')
YELLOW_FILL = PatternFill(start_color=PASTEL_YELLOW, end_color=PASTEL_YELLOW, fill_type='solid')
BLUE_FILL = PatternFill(start_color=PASTEL_BLUE, end_color=PASTEL_BLUE, fill_type='solid')
ORANGE_FILL = PatternFill(start_color=PASTEL_ORANGE, end_color=PASTEL_ORANGE, fill_type='solid')
HEADER_FILL = PatternFill(start_color=PASTEL_BLUE, end_color=PASTEL_BLUE, fill_type='solid')  # Light pastel, not dark

# === ALIGNMENTS ===
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
CENTER_ACROSS = Alignment(horizontal='centerContinuous', vertical='center')  # NOT merge & center
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

# === BORDERS ===
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Assumption cells: yellow fill + medium outside border (Wall St standard)
ASSUMPTION_BORDER = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)

# === CHART DIMENSIONS: 3" tall × 5" wide ===
CHART_HEIGHT_CM = 7.62  # 3 inches
CHART_WIDTH_CM = 12.7   # 5 inches

# === NUMBER FORMATS ===
# Currency: accounting format with red negatives
CURRENCY_FORMAT = '_($* #,##0.0_);[Red]_($* (#,##0.0);_($* "-"??_);_(@_)'
# Ratio: x suffix with red negatives
RATIO_FORMAT = '0.0"x";[Red](0.0"x");"-"'
# Percentage: with red negatives
PERCENT_FORMAT = '0.0%;[Red](0.0%);"-"'
# Standard number
NUMBER_FORMAT = '#,##0.0;[Red](#,##0.0);"-"'

# === COLUMN WIDTHS ===
NUMERIC_COL_WIDTH = 12  # Uniform width for numeric columns
LABEL_COL_WIDTH = 25    # Width for text/label columns
NAV_COL_WIDTH = 3       # Navigation column

# === ROW HEIGHTS ===
HEADER_ROW_HEIGHT = 30  # Taller for wrapped headers
DATA_ROW_HEIGHT = 14.5  # Excel default


def fix_em_dashes(text):
    """Replace em dashes and en dashes with hyphens"""
    if not isinstance(text, str):
        return text
    return text.replace('—', '-').replace('–', '-')


def fix_currency_notation(text):
    """Standardize currency: VNDk/mm/bn, US$k/mm/bn - never capitalized M or B"""
    if not isinstance(text, str):
        return text

    # Fix VND notation - lowercase mm/bn
    text = re.sub(r'VND\s*(\d+(?:[.,]\d+)?)\s*[Bb](?:illion)?', r'VND\1bn', text)
    text = re.sub(r'VND\s*(\d+(?:[.,]\d+)?)\s*[Mm](?:illion)?', r'VND\1mm', text)
    text = re.sub(r'VND\s*(\d+(?:[.,]\d+)?)\s*[Kk](?:ilo)?', r'VND\1k', text)

    # Fix standalone B/M/K after numbers to bn/mm/k
    text = re.sub(r'(\d)B\b', r'\1bn', text)
    text = re.sub(r'(\d)M\b', r'\1mm', text)
    text = re.sub(r'(\d)K\b', r'\1k', text)

    # Fix USD notation
    text = re.sub(r'\$(\d+(?:[.,]\d+)?)\s*[Bb](?:illion)?', r'US$\1bn', text)
    text = re.sub(r'\$(\d+(?:[.,]\d+)?)\s*[Mm](?:illion)?', r'US$\1mm', text)
    text = re.sub(r'\$(\d+(?:[.,]\d+)?)\s*[Kk]', r'US$\1k', text)
    text = re.sub(r'USD\s*(\d+(?:[.,]\d+)?)\s*[Bb]', r'US$\1bn', text)
    text = re.sub(r'USD\s*(\d+(?:[.,]\d+)?)\s*[Mm]', r'US$\1mm', text)
    text = re.sub(r'USD\s*(\d+(?:[.,]\d+)?)\s*[Kk]', r'US$\1k', text)

    return text


def to_title_case(text):
    """Convert ALL CAPS to Title Case"""
    if not isinstance(text, str):
        return text
    # Only convert if it's ALL CAPS and longer than 3 chars (preserve acronyms like USD, VND)
    if text.isupper() and len(text) > 3:
        return text.title()
    return text


def is_numeric_cell(value):
    """Check if cell contains numeric data"""
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        clean = value.replace(',', '').replace('$', '').replace('%', '').replace('(', '').replace(')', '').replace('VND', '').replace('US', '').strip()
        try:
            float(clean)
            return True
        except ValueError:
            return False
    return False


def is_percentage_value(value, number_format=None):
    """Check if cell is a percentage"""
    if number_format and '%' in str(number_format):
        return True
    if isinstance(value, str) and '%' in value:
        return True
    # Decimal between -1 and 1 (excluding 0) likely a percentage
    if isinstance(value, float) and -1 <= value <= 1 and value != 0 and value != 1 and value != -1:
        return True
    return False


def is_ratio_column(header_text):
    """Check if column header indicates ratio/multiple"""
    if not header_text:
        return False
    header = str(header_text).lower()
    ratio_indicators = ['p/e', 'ev/', 'p/b', 'p/s', 'pb', 'ps', 'pe', 'ratio', 'd/e', 'multiple', '/ebitda', '/rev', '/sales']
    return any(ind in header for ind in ratio_indicators)


def is_currency_column(header_text):
    """Check if column header indicates currency values"""
    if not header_text:
        return False
    header = str(header_text).lower()
    currency_indicators = ['us$', 'usd', 'vnd', '$', 'price', 'cap', 'revenue', 'ebitda', 'income', 'profit',
                           'debt', 'cash', 'assets', 'equity', 'ev ', 'auv', 'adtv']
    return any(ind in header for ind in currency_indicators)


def is_percentage_column(header_text):
    """Check if column header indicates percentage values"""
    if not header_text:
        return False
    header = str(header_text).lower()
    pct_indicators = ['%', 'margin', 'yield', 'roe', 'roa', 'roic', 'growth', 'cagr', 'pct', 'percent', 'franchise']
    return any(ind in header for ind in pct_indicators)


def get_cell_font_by_type(cell):
    """
    Wall Street color coding convention:
    - Blue (#0000FF): Hardcoded values (no formula)
    - Black (#000000): Formulas / calculated values
    - Green (#006400): References to other sheets in same workbook
    - Purple (#800080): References to other workbooks (external links)
    """
    # Check if cell has a formula
    if cell.data_type == 'f' or (hasattr(cell, 'value') and isinstance(cell.value, str) and str(cell.value).startswith('=')):
        formula = str(cell.value) if cell.value else ''

        # Check for external workbook reference: [Workbook.xlsx]Sheet!A1
        if '[' in formula and ']' in formula:
            return FONT_WORKBOOK_REF  # Purple

        # Check for sheet reference: Sheet!A1 or 'Sheet Name'!A1
        if '!' in formula:
            return FONT_SHEET_REF  # Green

        # Regular formula
        return FONT_FORMULA  # Black

    # No formula = hardcoded value
    return FONT_HARDCODE  # Blue


def is_assumption_cell(cell, sheet, header_row):
    """
    Detect assumption cells - typically:
    - Hardcoded numeric inputs
    - Often in specific rows/sections marked as assumptions
    - Usually blue (hardcoded) + need yellow fill + outside border

    Heuristic: Check if row label contains 'assumption' or cell is in assumption section
    """
    # Check if this row's label contains assumption-related keywords
    row_label = sheet.cell(row=cell.row, column=2).value
    if row_label and isinstance(row_label, str):
        label_lower = row_label.lower()
        if any(kw in label_lower for kw in ['assumption', 'input', 'driver', 'sensitivity', 'scenario']):
            # Only mark as assumption if it's a hardcoded numeric value
            if cell.data_type != 'f' and is_numeric_cell(cell.value):
                return True
    return False


def unmerge_all_cells(sheet):
    """Unmerge all merged cells - we use Center Across Selection instead"""
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_range in merged_ranges:
        sheet.unmerge_cells(str(merged_range))


def insert_navigation_column(sheet):
    """Insert Column A for navigation if data exists in column A"""
    # Check if column A has data
    col_a_has_data = any(sheet.cell(row=r, column=1).value for r in range(1, sheet.max_row + 1))

    if col_a_has_data:
        sheet.insert_cols(1)

    # Set Column A width narrow (about 3 characters)
    sheet.column_dimensions['A'].width = 3


def insert_empty_row1(sheet):
    """Insert empty Row 1 if data exists in row 1"""
    row1_has_data = any(sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1))

    if row1_has_data:
        sheet.insert_rows(1)


def detect_section_starts(sheet):
    """Detect section start rows (rows where column B has text but subsequent cells are mostly empty)"""
    section_rows = []

    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (row 1 is empty)
        cell_b = sheet.cell(row=row_idx, column=2)

        if cell_b.value and isinstance(cell_b.value, str):
            # Check if this looks like a section header
            # (text in B, mostly empty in C onwards, or bold formatting)
            cells_after_b = [sheet.cell(row=row_idx, column=c).value for c in range(3, min(8, sheet.max_column + 1))]
            empty_count = sum(1 for v in cells_after_b if not v)

            # If mostly empty after column B, likely a section header
            if empty_count >= len(cells_after_b) * 0.7:
                section_rows.append(row_idx)

    return section_rows


def add_navigation_markers(sheet):
    """Add Wingdings 'v' markers in Column A at section starts"""
    section_rows = detect_section_starts(sheet)

    # Add markers at section starts
    for row_idx in section_rows:
        cell = sheet.cell(row=row_idx, column=1)
        cell.value = 'v'  # Wingdings checkmark/arrow
        cell.font = NAV_FONT

    # Add end marker
    end_row = sheet.max_row + 2
    end_cell = sheet.cell(row=end_row, column=1)
    end_cell.value = 'v'
    end_cell.font = NAV_FONT

    label_cell = sheet.cell(row=end_row, column=2)
    label_cell.value = 'End Model'
    label_cell.font = LABEL_FONT


def normalize_row_heights(sheet, header_row=2):
    """Set uniform row heights - taller header, uniform data rows"""
    for row_idx in range(1, sheet.max_row + 1):
        if row_idx == header_row:
            sheet.row_dimensions[row_idx].height = HEADER_ROW_HEIGHT
        else:
            sheet.row_dimensions[row_idx].height = DATA_ROW_HEIGHT


def detect_header_row(sheet):
    """Detect the header row (usually row 2 since row 1 is empty)"""
    for row_idx in range(2, min(6, sheet.max_row + 1)):
        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(2, min(12, sheet.max_column + 1))]
        non_empty = [v for v in row_values if v]

        if len(non_empty) >= 2:
            # Check if mostly text (header)
            text_count = sum(1 for v in non_empty if isinstance(v, str) and not is_numeric_cell(v))
            if text_count >= len(non_empty) * 0.5:
                return row_idx
    return 2  # Default to row 2


def format_header_row(sheet, header_row):
    """Format header row - pastel fill, black text, wrap text for multi-line headers"""
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col)

        if col == 1:  # Navigation column
            continue

        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL  # Pastel blue, not dark
        cell.border = THIN_BORDER

        # Fix text content
        if cell.value:
            cell.value = to_title_case(fix_em_dashes(str(cell.value)))

        # Check if this column has numeric data below - if so, right-align header too
        has_numeric_below = False
        for check_row in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            check_cell = sheet.cell(row=check_row, column=col)
            if is_numeric_cell(check_cell.value):
                has_numeric_below = True
                break

        # Wrap text and vertical center for headers (allows multi-line)
        if has_numeric_below:
            cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def format_data_cells(sheet, header_row):
    """
    Format data cells with Wall Street color coding:
    - Blue: Hardcoded values
    - Black: Formulas
    - Green: References to other sheets
    - Purple: References to other workbooks
    - Yellow fill + outside border: Assumption cells
    """
    # First, determine column types from headers
    col_types = {}
    for col in range(1, sheet.max_column + 1):
        header_cell = sheet.cell(row=header_row, column=col)
        header_text = header_cell.value
        if is_ratio_column(header_text):
            col_types[col] = 'ratio'
        elif is_percentage_column(header_text):
            col_types[col] = 'percent'
        elif is_currency_column(header_text):
            col_types[col] = 'currency'
        else:
            col_types[col] = 'text'

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col)

            if col == 1:  # Navigation column - skip
                continue

            # Apply color-coded font based on cell type (Wall St convention)
            cell.font = get_cell_font_by_type(cell)
            cell.border = THIN_BORDER

            # Check if this is an assumption cell - apply yellow fill + outside border
            if is_assumption_cell(cell, sheet, header_row):
                cell.fill = YELLOW_FILL
                cell.border = ASSUMPTION_BORDER

            # Fix text content
            if isinstance(cell.value, str) and not str(cell.value).startswith('='):
                cell.value = fix_em_dashes(fix_currency_notation(to_title_case(cell.value)))

            # Alignment and number format based on content type
            if is_numeric_cell(cell.value):
                # Wrap text + right align for data rows
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

                # Apply number format based on column type
                col_type = col_types.get(col, 'text')
                if col_type == 'ratio':
                    cell.number_format = RATIO_FORMAT
                elif col_type == 'percent':
                    cell.number_format = PERCENT_FORMAT
                elif col_type == 'currency':
                    cell.number_format = CURRENCY_FORMAT
                else:
                    # Check if it looks like a percentage value anyway
                    if is_percentage_value(cell.value, cell.number_format):
                        cell.number_format = PERCENT_FORMAT
                    else:
                        cell.number_format = NUMBER_FORMAT
            else:
                # Wrap text + left align for text
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def unhide_rows_columns(sheet):
    """Unhide all rows and columns - never hide, use grouping only if needed"""
    for row_idx in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_idx].hidden = False

    for col_idx in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].hidden = False


def detect_data_start_column(sheet, header_row):
    """Detect first column with actual data (not navigation/labels)"""
    for col in range(1, sheet.max_column + 1):
        # Check a few rows below header for numeric data
        for row in range(header_row + 1, min(header_row + 5, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col)
            if is_numeric_cell(cell.value):
                return col
    return 2  # Default to column B if no numeric data found


def setup_freeze_panes(sheet, header_row):
    """Freeze title rows + column prior to data"""
    # Find first data column, freeze the column before it
    data_col = detect_data_start_column(sheet, header_row)
    freeze_col = get_column_letter(data_col)
    freeze_cell = f'{freeze_col}{header_row + 1}'
    sheet.freeze_panes = freeze_cell


def set_column_widths(sheet, header_row):
    """Set uniform widths - numeric columns similar width, label columns wider"""
    # Column A (navigation) always narrow
    sheet.column_dimensions['A'].width = NAV_COL_WIDTH

    for col in range(2, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        header_cell = sheet.cell(row=header_row, column=col)
        header_text = header_cell.value

        # Check if column has numeric data
        has_numeric = False
        for check_row in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            if is_numeric_cell(sheet.cell(row=check_row, column=col).value):
                has_numeric = True
                break

        if has_numeric:
            # All numeric columns get uniform width
            sheet.column_dimensions[col_letter].width = NUMERIC_COL_WIDTH
        else:
            # Text/label columns get wider width
            sheet.column_dimensions[col_letter].width = LABEL_COL_WIDTH


def format_charts(sheet):
    """Format all charts: 3"×5", Aptos fonts, no gridlines, show data labels"""
    for chart in sheet._charts:
        # Set dimensions: 3 inches tall × 5 inches wide
        chart.height = CHART_HEIGHT_CM
        chart.width = CHART_WIDTH_CM

        # Title formatting: 12pt bold Aptos
        if chart.title:
            # Can't directly set font on title in openpyxl, but we ensure title case
            if hasattr(chart.title, 'text') and chart.title.text:
                chart.title.text = to_title_case(fix_em_dashes(chart.title.text))

        # Remove major gridlines
        if hasattr(chart, 'y_axis') and chart.y_axis:
            chart.y_axis.majorGridlines = None
            chart.y_axis.minorGridlines = None
        if hasattr(chart, 'x_axis') and chart.x_axis:
            chart.x_axis.majorGridlines = None
            chart.x_axis.minorGridlines = None

        # Add data labels to all series
        if hasattr(chart, 'series'):
            for series in chart.series:
                series.labels = DataLabelList()
                series.labels.showVal = True
                series.labels.showPercent = False  # Show values, format as % if needed


def format_worksheet(sheet):
    """Format a single worksheet with all standards"""
    if sheet.max_row == 0 or sheet.max_column == 0:
        return

    print(f"  Processing sheet: {sheet.title}")

    # Step 1: Unmerge all cells (we use Center Across Selection instead)
    unmerge_all_cells(sheet)

    # Step 2: Insert empty Row 1 if needed
    insert_empty_row1(sheet)

    # Step 3: Insert navigation Column A if needed
    insert_navigation_column(sheet)

    # Step 4: Unhide all rows and columns
    unhide_rows_columns(sheet)

    # Step 5: Detect header row
    header_row = detect_header_row(sheet)

    # Step 6: Set uniform row heights (taller header, uniform data)
    normalize_row_heights(sheet, header_row)

    # Step 7: Set column widths (uniform numeric, wider labels)
    set_column_widths(sheet, header_row)

    # Step 8: Format header row
    format_header_row(sheet, header_row)

    # Step 9: Format data cells with proper number formats
    format_data_cells(sheet, header_row)

    # Step 10: Add navigation markers in Column A
    add_navigation_markers(sheet)

    # Step 11: Format charts
    format_charts(sheet)

    # Step 12: Setup freeze panes
    setup_freeze_panes(sheet, header_row)


def format_workbook(filepath):
    """Main formatting function"""
    print(f"Loading: {filepath}")
    wb = load_workbook(filepath)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        format_worksheet(sheet)

    # Save
    wb.save(filepath)
    print(f"Saved: {filepath}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: format_excel.py <filepath>")
        sys.exit(1)

    filepath = sys.argv[1]
    if not Path(filepath).exists():
        print(f"Error: File not found: {filepath}")
        sys.exit(1)

    format_workbook(filepath)
